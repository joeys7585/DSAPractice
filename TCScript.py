# daily_run_refactored.py (fully extended version)

import socket
import os
import struct
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from Cryptodome.Cipher import AES
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import time

# Global store for test results
test_results = {
    "non_persistent_orderno": [],
    "user_id": [],
    "message_seq_no": [],
    "template_id_response": [],
    "testcase_names": [],
    "act_responses": [],
    "expected_responses": [],
    "comments": []
}

# 1. Environment and Configuration
def get_env_config(env_name):
    env = env_name.upper()
    ip_map = {
        "MAT": ("192.168.189.35", 19106),
        "UAT": ("192.168.190.40", 19306),
        "NFR": ("192.168.80.48", 19006),
        "UAT2": ("192.168.190.201", 19406),
        "MOCKDR": ("192.168.73.48", 19106),
        "MOCKDC": ("192.168.63.48", 19106)
    }
    key_iv_map = {
        "UAT2": (b"Mcx@5432100000000000000000000000", b"Mcx@54321000")
    }
    return {
        "ip": ip_map[env][0],
        "port": ip_map[env][1],
        "key": key_iv_map.get(env, (None, None))[0],
        "iv": key_iv_map.get(env, (None, None))[1]
    }

# 2. Socket Connection
def connect_to_socket(ip, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.connect((ip, port))
    return sock

# 3. AES-GCM Encryption/Decryption
def init_gcm_cipher(key, iv):
    cipher = Cipher(algorithms.AES(key), modes.GCM(iv), backend=default_backend())
    return cipher.encryptor(), cipher.decryptor()

# 4. Payload Handling
def create_payload(field_values, fmt):
    result = []
    for val in field_values:
        if isinstance(val, str):
            result.append(val.replace("\\0", "\0").encode("utf-8"))
        else:
            result.append(val)
    return struct.pack(fmt, *result)

# 5. Excel Handling
def load_excel_sheets(filepath, sheet_map):
    dfs = {}
    for sheet in sheet_map:
        dfs[sheet] = pd.read_excel(filepath, sheet_name=sheet, engine='openpyxl')
    return dfs

# 6. Order Execution Logic with Parsing, Expected Comparison, Pass/Fail
def execute_order_flow(sock, dataframes, key, iv):
    config_df = dataframes["Configuration"]
    order_df = dataframes["Order_data"]

    cipher = Cipher(algorithms.AES(key), modes.GCM(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    decryptor = cipher.decryptor()

    for i, row in config_df.iterrows():
        if str(row.get("status", "N")).upper() != "Y":
            continue

        order_type = str(row.get("order_type", "")).upper()
        testcase = row.get("Test_case_name", f"TC_{i+1}")
        row_num = int(row.get("Order_no(Row_number)", 0))
        expected_template = str(row.get("Expected_TemplateID", "")).strip()

        print(f"[TEST] Executing: {testcase} [{order_type}]")

        order_data = order_df.loc[row_num]
        test_results["testcase_names"].append(testcase)
        test_results["user_id"].append(order_data[5])
        test_results["message_seq_no"].append(order_data[4])
        test_results["expected_responses"].append(expected_template)

        # Message creation based on order type
        if order_type == "NEW":
            header_fields = order_data.loc["BodyLen":"Price"].iloc[0:6]
            body_fields = order_data.loc["Price":"FreeText3"]
            header_fields[4] = order_data[4]  # message_seq_no
            header = create_payload(header_fields, "IH8s2sII")
            body = create_payload(body_fields, "qqQQQQQQIiqqIiIIH5s7s9sBBBBBBBBBBB2sc2s1s20s12s12s12s")

        elif order_type == "MODIFY":
            header_fields = order_data.loc["MBodyLen":"SenderSubID"]
            body_fields = order_data.loc["MOrderID":"MPad4"]
            body_fields[0] = int(test_results["non_persistent_orderno"][row_num])
            header = create_payload(header_fields, "IH8s2sII")
            body = create_payload(body_fields, "QQQqqQQQQIiqqQQIiIIIH5s7s9sBBBBBBBBBB1s2s1s2s1s20s12s12s12s4s")

        elif order_type == "CANCEL":
            header_fields = order_data.loc["CBodyLen":"COrderID"].iloc[0:6]
            body_fields = order_data.loc["COrderID":"CPad4"]
            body_fields[0] = int(test_results["non_persistent_orderno"][row_num])
            header = create_payload(header_fields, "IH8s2sII")
            body = create_payload(body_fields, "QQQQQQiiIII4s")

        else:
            print(f"[WARN] Unsupported order type: {order_type}")
            continue

        encrypted = encryptor.update(body)
        message = header + encrypted

        sock.sendall(message)
        time.sleep(1)

        try:
            header_resp = sock.recv(8)
            template_id = struct.unpack_from("IH2s", header_resp)[1]
            test_results["template_id_response"].append(template_id)

            payload = sock.recv(1024)
            decrypted = decryptor.update(payload)

            # Parse non-persistent order number from decrypted payload (first 8 bytes as example)
            order_no = struct.unpack_from("Q", decrypted)[0]
            test_results["non_persistent_orderno"].append(order_no)

            # Compare with expected response
            status = "PASS" if template_id.decode().strip() == expected_template else "FAIL"
            test_results["act_responses"].append(f"{template_id.decode()} - {order_no}")
            test_results["comments"].append(status)

        except Exception as e:
            print(f"[ERROR] Socket/Decode issue: {e}")
            test_results["template_id_response"].append("ERROR")
            test_results["non_persistent_orderno"].append("-")
            test_results["act_responses"].append("ERROR")
            test_results["comments"].append("FAIL")

# 7. Report Generation with Color Coding
def generate_excel_report(filepath):
    print(f"[INFO] Generating report at {filepath}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["Sr_No", "TestCase_name", "userid", "orderno", "TemplateID", "Responses", "Expected", "Comment"]
    ws.append(headers)

    fill_header = PatternFill(patternType='solid', fgColor='00FFFFCC')
    fill_pass = PatternFill(patternType='solid', fgColor='0099FF99')  # Light green
    fill_fail = PatternFill(patternType='solid', fgColor='00FF9999')  # Light red

    for cell in ws[1]:
        cell.fill = fill_header

    for i, testcase in enumerate(test_results["testcase_names"]):
        row = [
            i + 1,
            testcase,
            test_results["user_id"][i],
            test_results["non_persistent_orderno"][i],
            test_results["template_id_response"][i],
            test_results["act_responses"][i],
            test_results["expected_responses"][i],
            test_results["comments"][i]
        ]
        ws.append(row)
        fill = fill_pass if test_results["comments"][i] == "PASS" else fill_fail
        for cell in ws[i + 2]:
            cell.fill = fill

    wb.save(filepath)
    print("[INFO] Report generation complete.")

# 8. Main Execution (Flow Controller)
def main(env_name="UAT"):
    config = get_env_config(env_name)
    current_dir = os.getcwd()
    excel_path = os.path.join(current_dir, "ETI_UAT_Algo_10003.xlsx")

    print(f"\n[INFO] Running in {env_name.upper()} mode")
    sock = connect_to_socket(config["ip"], config["port"])

    sheet_map = ["All_login", "Order_data", "Configuration"]
    dataframes = load_excel_sheets(excel_path, sheet_map)

    if config["key"] and config["iv"]:
        encryptor, decryptor = init_gcm_cipher(config["key"], config["iv"])
    else:
        encryptor = decryptor = None

    try:
        execute_order_flow(sock, dataframes, config["key"], config["iv"])
    finally:
        sock.close()
        report_path = os.path.join(current_dir, "Report", f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        generate_excel_report(report_path)

if __name__ == "__main__":
    import sys
    env_name = sys.argv[1] if len(sys.argv) > 1 else "UAT"
    main(env_name)
